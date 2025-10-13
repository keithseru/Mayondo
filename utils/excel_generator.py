from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference, PieChart, LineChart
from collections import Counter
from datetime import datetime

class ExcelReportGenerator:
    '''Base class for generating Excel reports'''
    
    def __init__(self, title):
        self.title = title
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remove the default sheet
        
        # Define styles
        self.header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='2C5F2D', end_color='2C5F2D', fill_type='solid')
        self.title_font = Font(name='Arial', size=18, bold=True, color='2C5F2D')
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_sheet(self, name):
        '''Create a new worksheet'''
        sheet = self.workbook.create_sheet(title=name)
        return sheet
    
    def add_title(self, sheet, title, row=1):
        '''Add a title to the sheet'''
        sheet.merge_cells(f'A{row}:E{row}')
        cell = sheet[f'A{row}']
        cell.value = title
        cell.font = self.title_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.row_dimensions[row].height = 30
        
    def add_date(self, sheet, row=2):
        '''Add generation date'''
        sheet.merge_cells(f'A{row}:E{row}')
        cell = sheet[f'A{row}']
        cell.value = f"Generated: {datetime.now().strftime('%d %B, %Y at %I:%M %p')}"
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(size=10, italic=True)
        
    def style_header_row(self, sheet, row, columns):
        '''Style the header rows'''
        for col_num, _ in enumerate(columns, 1):
            cell = sheet.cell(row=row, column=col_num)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        sheet.row_dimensions[row].height = 20
        
    def add_summary_section(self, sheet, summary_data, start_row):
        '''Add summary statistics'''
        sheet.merge_cells(f'A{start_row}:B{start_row}')
        cell = sheet[f'A{start_row}']
        cell.value = 'Summary'
        cell.font = Font(size=12, bold=True)
        
        row = start_row + 1
        for label, value in summary_data.items():
            sheet[f'A{row}'] = label
            sheet[f'B{row}'] = value
            sheet[f'A{row}'].font = Font(bold=True)
            sheet[f'A{row}'].border = self.border
            sheet[f'B{row}'].border = self.border
            sheet[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
            sheet[f'B{row}'].alignment = Alignment(horizontal='right', vertical='center')
            row += 1
        return row + 1
    
    def auto_adjust_columns(self, sheet):
        '''Auto-adjust column widths - Fixed for merged cells'''
        from openpyxl.cell import MergedCell
        
        for column in sheet.columns:
            max_length = 0
            column_letter = None
            
            for cell in column:
                # Skip merged cells
                if isinstance(cell, MergedCell):
                    continue
                
                # Get column letter from first non-merged cell
                if column_letter is None:
                    column_letter = cell.column_letter
                
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Only adjust if we found a valid column letter
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
    def save(self, buffer):
        '''Save the workbook to buffer'''
        self.workbook.save(buffer)
        return buffer
    
class SalesReportExcel(ExcelReportGenerator):
    '''Sales report Excel generator'''
    
    def __init__(self, date_from, date_to, sales_data, summary):
        title = f"Sales Report: {date_from.strftime('%d %B, %Y')} to {date_to.strftime('%d %B, %Y')}"
        super().__init__(title)
        self.date_from = date_from
        self.date_to = date_to
        self.sales_data = sales_data
        self.summary = summary
        
    def build(self):
        '''Build the sales report'''
        sheet = self.create_sheet('Sales Report')
        
        # Title
        self.add_title(sheet, self.title)
        self.add_date(sheet, row=2)
        
        # Summary Section - Updated to include all status counts
        summary_data = {
            'Total Sales': self.summary['total_sales'],
            'Total Revenue': f"UGX {self.summary['total_revenue']:,}",
            'Average Sale': f"UGX {self.summary['average_sale']:,.0f}",
            'Total Customers': self.summary['total_customers'],
            'Completed Sales': self.summary.get('completed_sales', 0),
            'Pending Sales': self.summary.get('pending_sales', 0),
            'Cancelled Sales': self.summary.get('cancelled_sales', 0),
        }
        data_start_row = self.add_summary_section(sheet, summary_data, start_row=4)
        
        # Sales data table
        headers = ['Date', 'Customer', "Items", 'Amount(UGX)', 'Status']
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=data_start_row, column=col_num, value=header)
        self.style_header_row(sheet, data_start_row, headers)
        
        # Add data
        row = data_start_row + 1
        for sale in self.sales_data:
            sheet[f'A{row}'] = sale['date'].strftime('%m/%d/%Y')
            sheet[f'B{row}'] = sale['customer']
            sheet[f'C{row}'] = sale['items']
            sheet[f'D{row}'] = sale['amount']
            sheet[f'E{row}'] = sale['status']
            
            # Apply borders
            for col in range(1, 6):
                sheet.cell(row=row, column=col).border = self.border
                
            row += 1
            
        # Auto-adjust columns
        self.auto_adjust_columns(sheet)
        
        # Add charts if we have data
        if len(self.sales_data) > 0:
            self.add_sales_chart(sheet, data_start_row, row - 1)
            self.add_revenue_line_chart(sheet, data_start_row, row - 1)
            self.add_status_pie_chart()
        
        return self
    
    def add_sales_chart(self, sheet, start_row, end_row):
        '''Add a sales chart with improved visuals'''
        chart_sheet = self.create_sheet('Sales Chart')
        
        # Add sheet title
        chart_sheet.merge_cells('A1:F1')
        title_cell = chart_sheet['A1']
        title_cell.value = "Sales by Date - Bar Chart"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        chart_sheet.row_dimensions[1].height = 25
        
        # Create bar chart
        chart = BarChart()
        chart.type = "col"  # Column chart (vertical bars)
        chart.style = 10  # Professional color scheme
        chart.title = "Sales Revenue by Date"
        chart.y_axis.title = 'Amount (UGX)'
        chart.x_axis.title = 'Transaction Date'
        
        # Set data
        data = Reference(sheet, min_col=4, min_row=start_row, max_row=end_row)
        categories = Reference(sheet, min_col=1, min_row=start_row+1, max_row=end_row)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        # Improve chart dimensions and spacing
        chart.width = 20  # Wider chart
        chart.height = 12  # Taller chart
        
        # Rotate x-axis labels for better readability
        chart.x_axis.tickLblPos = "low"
        
        # Add chart with spacing from title
        chart_sheet.add_chart(chart, 'A3')
    
    def add_revenue_line_chart(self, sheet, start_row, end_row):
        '''Add revenue trend line chart with improved visuals'''
        chart_sheet = self.create_sheet("Revenue Trend")
        
        # Add sheet title
        chart_sheet.merge_cells('A1:F1')
        title_cell = chart_sheet['A1']
        title_cell.value = "Revenue Trend Over Time"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        chart_sheet.row_dimensions[1].height = 25
        
        # Create a line chart
        chart = LineChart()
        chart.title = 'Cumulative Revenue Trend'
        chart.style = 12  # Professional style
        chart.y_axis.title = "Revenue (UGX)"
        chart.x_axis.title = "Date"
        
        data = Reference(sheet, min_col=4, min_row=start_row, max_row=end_row)
        categories = Reference(sheet, min_col=1, min_row=start_row + 1, max_row=end_row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        # Improve chart dimensions and spacing
        chart.width = 20  # Wider chart
        chart.height = 12  # Taller chart
        
        # Style the line - do this after adding data
        try:
            from openpyxl.chart.marker import Marker
            from openpyxl.drawing.line import LineProperties
            from openpyxl.drawing.fill import SolidColorFillProperties, ColorChoice
            
            if chart.series and len(chart.series) > 0:
                series = chart.series[0]
                # Add markers
                series.marker = Marker('circle')
                series.marker.size = 5
                
                # Style the line
                line_props = LineProperties(solidFill=SolidColorFillProperties(ColorChoice(srgbClr="2C5F2D")))
                line_props.width = 25000
                series.graphicalProperties.line = line_props
        except Exception as e:
            # If styling fails, continue without it
            pass
        
        # Rotate x-axis labels for better readability
        chart.x_axis.tickLblPos = "low"

        # Add chart with spacing from title
        chart_sheet.add_chart(chart, "A3")

    def add_status_pie_chart(self):
        '''Add status breakdown pie chart - FIXED positioning and sizing'''
        chart_sheet = self.create_sheet("Status Breakdown")

        # Count statuses
        status_counts = Counter(sale['status'] for sale in self.sales_data)

        # Add title with proper spacing
        chart_sheet.merge_cells('A1:B1')
        title_cell = chart_sheet['A1']
        title_cell.value = "Sales Status Breakdown"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Add spacing
        chart_sheet.row_dimensions[1].height = 25
        chart_sheet.row_dimensions[2].height = 10

        # Write data to sheet starting from row 3
        chart_sheet['A3'] = "Status"
        chart_sheet['B3'] = "Count"
        chart_sheet['A3'].font = Font(bold=True)
        chart_sheet['B3'].font = Font(bold=True)
        
        # Adjust column widths for better visibility
        chart_sheet.column_dimensions['A'].width = 20
        chart_sheet.column_dimensions['B'].width = 15
        
        # Add borders to header row
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        chart_sheet['A3'].border = thin_border
        chart_sheet['B3'].border = thin_border
        
        # Add header background color
        header_fill = PatternFill(start_color='2C5F2D', end_color='2C5F2D', fill_type='solid')
        chart_sheet['A3'].fill = header_fill
        chart_sheet['B3'].fill = header_fill
        chart_sheet['A3'].font = Font(bold=True, color='FFFFFF')
        chart_sheet['B3'].font = Font(bold=True, color='FFFFFF')
        
        row = 4
        for status, count in status_counts.items():
            chart_sheet[f'A{row}'] = status
            chart_sheet[f'B{row}'] = count
            # Add borders to data cells
            chart_sheet[f'A{row}'].border = thin_border
            chart_sheet[f'B{row}'].border = thin_border
            row += 1

        # Create pie chart WITHOUT internal title
        chart = PieChart()
        chart.title = None  # Remove the overlapping title
        data = Reference(chart_sheet, min_col=2, min_row=3, max_row=row - 1)
        labels = Reference(chart_sheet, min_col=1, min_row=4, max_row=row - 1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        
        # Set chart dimensions - made wider to accommodate legend
        chart.width = 18  # Increased width
        chart.height = 12
        
        # Position legend to the right to avoid overlap
        from openpyxl.chart.legend import Legend
        chart.legend = Legend()
        chart.legend.position = 'r'  # Right position
        chart.legend.overlay = False  # Don't overlay on chart

        # Position chart to the right with good spacing
        chart_sheet.add_chart(chart, "D3")  # Aligned with table header

class InventoryReportExcel(ExcelReportGenerator):
    """Inventory report Excel generator"""
    
    def __init__(self, inventory_data, summary):
        super().__init__("Inventory Status Report")
        self.inventory_data = inventory_data
        self.summary = summary
    
    def build(self):
        """Build the inventory report"""
        # Create inventory sheet
        sheet = self.create_sheet("Inventory Report")

        # Title and date
        self.add_title(sheet, self.title)
        self.add_date(sheet, row=2)

        # Summary section
        summary_data = {
            'Total Products:': self.summary['total_items'],
            'Total Variants:': self.summary['total_variants'],
            'Low Stock Items:': self.summary['low_stock'],
            'Out of Stock:': self.summary['out_of_stock'],
            'Total Inventory Value:': f"UGX {self.summary['total_value']:,.0f}",
        }
        data_start_row = self.add_summary_section(sheet, summary_data, start_row=4)

        # Table headers
        headers = ['Product', 'Variant', 'Current Stock', 'Reorder Level', 'Status']
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=data_start_row, column=col_num, value=header)
        self.style_header_row(sheet, data_start_row, headers)

        # Table data
        row = data_start_row + 1
        for item in self.inventory_data['low_stock']:
            status = "OUT OF STOCK" if item['stock'] == 0 else "LOW STOCK"

            sheet[f'A{row}'] = item['product']
            sheet[f'B{row}'] = item['variant']
            sheet[f'C{row}'] = item['stock']
            sheet[f'D{row}'] = item['reorder_level']
            sheet[f'E{row}'] = status

            # Color-code status
            status_cell = sheet[f'E{row}']
            if status == "OUT OF STOCK":
                status_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                status_cell.font = Font(color='FFFFFF', bold=True)
            else:
                status_cell.fill = PatternFill(start_color='FFC107', end_color='FFC107', fill_type='solid')

            # Highlight critical stock
            if item['stock'] < item['reorder_level']:
                sheet[f'C{row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            # Apply borders and alignment
            for col in range(1, 6):
                cell = sheet.cell(row=row, column=col)
                cell.border = self.border
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Center-align numeric columns
            sheet.cell(row=row, column=3).alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(row=row, column=4).alignment = Alignment(horizontal='center', vertical='center')

            row += 1

        # Auto-adjust columns
        self.auto_adjust_columns(sheet)

        # Add pie chart for stock status breakdown
        if self.summary['low_stock'] or self.summary['out_of_stock']:
            chart_sheet = self.create_sheet("Stock Status Chart")
            chart_sheet['A1'] = "Status"
            chart_sheet['B1'] = "Count"
            chart_sheet['A2'] = "LOW STOCK"
            chart_sheet['B2'] = self.summary['low_stock']
            chart_sheet['A3'] = "OUT OF STOCK"
            chart_sheet['B3'] = self.summary['out_of_stock']

            chart = PieChart()
            chart.title = "Stock Status Breakdown"
            data = Reference(chart_sheet, min_col=2, min_row=1, max_row=3)
            labels = Reference(chart_sheet, min_col=1, min_row=2, max_row=3)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            chart_sheet.add_chart(chart, "D2")

        return self